from openpyxl import load_workbook
from slugify import slugify


def list_of_dicts_from_excel(excel_filepath: str,
                             ingest_job_ignore_column: int = None,
                             worksheet_index: int = 0,
                             key_rows: object = None, first_item_row: int = 1,
                             ingest_job_id_column: int = 1):
    """
    Opens an Excel file and returns a list of dicts, where each dict represents details of an asset.
    Each row in the Excel file is assumed to either be a field name or details of an asset.
    Field names from multiple rows can be concatenated to create the dict keys.

    :param ingest_job_ignore_column: the column number to use as a job ignore flag
    :param ingest_job_id_column: the column number to use as job_id field
    :param excel_filepath: path to the Excel file
    :param worksheet_index: index of the sheet within the Excel file (zero-based)
    :param key_rows: list of row numbers to use when creating dict keys (one-based)
    :param first_item_row: the row number of the first asset (one-based)
    :return: a list of dicts
    """
    if key_rows is None:
        key_rows = [1]
    wb = load_workbook(excel_filepath, read_only=True, data_only=True)
    sheet = wb.worksheets[worksheet_index]

    keys = [
        slugify('-'.join(
            sheet.cell(j, i).value
            for j in key_rows
            if sheet.cell(j, i).value is not None
        )).replace('-', '_')
        for i in range(1, sheet.max_column+1)
    ]

    items = []
    for i in range(first_item_row, sheet.max_row+1):
        item = {keys[j - 1]: sheet.cell(i, j).value for j in range(1, len(keys)+1)}
        items.append(item)

    # close workbook
    wb.close()

    # add extra fields to items
    for item in items:
        item['ingest_job_id'] = item[keys[ingest_job_id_column-1]]
        item['ingest_job_ignore'] = True if ingest_job_ignore_column is not None and item[keys[ingest_job_ignore_column-1]] is not None else False

    return items
